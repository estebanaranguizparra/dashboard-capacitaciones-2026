#!/usr/bin/env python3
"""
Transforma el "Libro de Clases" (export BUK, formato ancho) en formato largo
(una fila por alumno-curso) y calcula prioridad / plan de accion para cada
capacitacion pendiente o en progreso.

Uso:
    python3 transform.py <ruta_excel_origen> <carpeta_salida>
"""
import sys
import csv
import json
from datetime import datetime, date
import openpyxl

BASE_COLS = [
    "id_alumno",
    "nombre_alumno",
    "estado_empleado",
    "fecha_ingreso",
    "seccion",
    "sucursal",
]
BLOCK_FIELDS = ["curso", "fecha_inicio", "fecha_termino", "estado_curso", "nota"]

PRIORIDAD_ALTA_DIAS = 0   # vencido (fecha_termino ya paso) y no completado
PRIORIDAD_MEDIA_DIAS = 15  # vence en los proximos 15 dias


def parse_fecha(v):
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    try:
        return datetime.strptime(str(v).strip(), "%d/%m/%Y").date()
    except ValueError:
        return None


def clasificar_prioridad(estado_curso, fecha_termino, hoy):
    """Devuelve (prioridad, accion) para una capacitacion no completada."""
    if estado_curso == "Completado":
        return "Sin accion", "Capacitacion completada."
    if fecha_termino is None:
        if estado_curso == "En progreso":
            return "Media", "Curso en progreso sin fecha de termino registrada: verificar avance."
        return "Media", "Curso pendiente sin fecha de termino registrada: agendar inicio."

    dias = (fecha_termino - hoy).days
    if dias < PRIORIDAD_ALTA_DIAS:
        return "Alta", f"Vencido hace {abs(dias)} dia(s). Contactar al colaborador y a su jefatura de forma inmediata."
    if dias <= PRIORIDAD_MEDIA_DIAS:
        return "Media", f"Vence en {dias} dia(s). Enviar recordatorio y asegurar cupo/acceso al curso."
    return "Baja", f"Vence en {dias} dia(s). Monitorear avance regular."


def main():
    if len(sys.argv) != 3:
        print("Uso: python3 transform.py <excel_origen> <carpeta_salida>")
        sys.exit(1)

    src, outdir = sys.argv[1], sys.argv[2]
    hoy = date.today()

    print(f"Cargando {src} ...")
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    max_row, max_col = ws.max_row, ws.max_column
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    print(f"Filas: {max_row}, columnas: {max_col}")

    n_blocks = (max_col - 6) // 5

    long_rows = []
    for r in range(2, max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        id_alumno = row_vals[0]
        if id_alumno is None or str(id_alumno).strip() == "":
            continue
        nombre_alumno = row_vals[1]
        estado_empleado = row_vals[2]
        fecha_ingreso = parse_fecha(row_vals[3])
        seccion = row_vals[4]
        sucursal = row_vals[5]

        for b in range(n_blocks):
            base = 6 + b * 5
            curso = row_vals[base]
            if curso is None or str(curso).strip() == "":
                continue
            fecha_inicio = parse_fecha(row_vals[base + 1])
            fecha_termino = parse_fecha(row_vals[base + 2])
            estado_curso = row_vals[base + 3] or "Pendiente"
            nota = row_vals[base + 4]

            prioridad, accion = clasificar_prioridad(estado_curso, fecha_termino, hoy)

            long_rows.append(
                {
                    "id_alumno": str(id_alumno).strip(),
                    "nombre_alumno": nombre_alumno,
                    "estado_empleado": estado_empleado,
                    "fecha_ingreso": fecha_ingreso.isoformat() if fecha_ingreso else "",
                    "seccion": seccion,
                    "sucursal": sucursal,
                    "curso": curso,
                    "fecha_inicio": fecha_inicio.isoformat() if fecha_inicio else "",
                    "fecha_termino": fecha_termino.isoformat() if fecha_termino else "",
                    "estado_curso": estado_curso,
                    "nota": nota if nota not in (None, "") else "",
                    "prioridad": prioridad,
                    "accion_sugerida": accion,
                }
            )

    print(f"Registros largos generados: {len(long_rows)}")

    fieldnames = [
        "id_alumno", "nombre_alumno", "estado_empleado", "fecha_ingreso",
        "seccion", "sucursal", "curso", "fecha_inicio", "fecha_termino",
        "estado_curso", "nota", "prioridad", "accion_sugerida",
    ]

    csv_path = f"{outdir}/capacitaciones_long.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(long_rows)
    print(f"CSV escrito: {csv_path}")

    # JSON normalizado y compacto para el dashboard: la prioridad y la accion
    # sugerida se recalculan en el navegador con la fecha real de visita
    # (no con la fecha de generacion del snapshot), asi nunca quedan obsoletas.
    alumno_index = {}
    alumnos = []
    curso_index = {}
    cursos = []
    capacitaciones = []

    for row in long_rows:
        aid = row["id_alumno"]
        if aid not in alumno_index:
            alumno_index[aid] = len(alumnos)
            alumnos.append(
                {
                    "id": aid,
                    "nombre": row["nombre_alumno"],
                    "estado_empleado": row["estado_empleado"],
                    "fecha_ingreso": row["fecha_ingreso"],
                    "seccion": row["seccion"],
                    "sucursal": row["sucursal"],
                }
            )
        curso = row["curso"]
        if curso not in curso_index:
            curso_index[curso] = len(cursos)
            cursos.append(curso)

        capacitaciones.append(
            [
                alumno_index[aid],
                curso_index[curso],
                row["fecha_inicio"],
                row["fecha_termino"],
                row["estado_curso"],
                row["nota"],
            ]
        )

    json_path = f"{outdir}/capacitaciones.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "generado": hoy.isoformat(),
                "columnas_capacitacion": [
                    "alumno_idx", "curso_idx", "fecha_inicio",
                    "fecha_termino", "estado_curso", "nota",
                ],
                "alumnos": alumnos,
                "cursos": cursos,
                "capacitaciones": capacitaciones,
            },
            f,
            ensure_ascii=False,
            separators=(",", ":"),
        )
    print(f"JSON escrito: {json_path}")


if __name__ == "__main__":
    main()
